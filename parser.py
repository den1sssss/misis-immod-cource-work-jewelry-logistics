from openpyxl import load_workbook

workbook = load_workbook('director_list.xlsx')

sheet = workbook.active

my_dict = {}

for row in sheet.iter_rows(values_only=True):
    key = str(row[-1])
    # values = [str(value) if True else datetime.strptime(str(value), '%d.%m.%Y').strftime('%d:%m:%Y') for idx, value
    #           in enumerate(row[:-1])] // ya ebal etu datu ebuchuyu
    values = [str(value) for idx, value in enumerate(row[:-1])]
    if key in my_dict:
        my_dict[key].append(values)
    else:
        my_dict[key] = [values]

for key, values_list in my_dict.items():
    print(key)
    for values in values_list:
        print(values)
